import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill

def is_odfpy_installed():
    try:
        import odf
        return True
    except ImportError:
        return False

def compare_tables(table1, table2, key_column):
    table1[key_column] = table1[key_column].astype(str).str.strip()
    table2[key_column] = table2[key_column].astype(str).str.strip()
    result = table1.copy()
    result['Совпадение в таблице 2'] = result[key_column].isin(table2[key_column])
    for column in table1.columns:
        if column == key_column:
            continue
        presence_column = f'{column} заполнено в таблице 2'
        result[presence_column] = result[key_column].map(
            lambda key: pd.notna(
                table2.loc[table2[key_column] == key, column].values[0]
            ) if key in table2[key_column].values and column in table2.columns else False
        )
    return result

def generate_excel_with_colors(df):
    wb = Workbook()
    ws = wb.active
    ws.title = "Сравнение"

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    headers = list(df.columns)
    ws.append(headers)

    for index, row in df.iterrows():
        ws.append(row.tolist())

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            col_name = headers[cell.col_idx - 1]
            if "Совпадение" in col_name:
                cell.fill = green_fill if cell.value else gray_fill
            elif "заполнено" in col_name:
                cell.fill = green_fill if cell.value else red_fill

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def main():
    st.title("🔍 Сравнение таблиц с цветной выгрузкой в Excel")

    odfpy_available = is_odfpy_installed()

    uploaded_file1 = st.file_uploader("Загрузите первую таблицу (шаблон)", type=["xlsx", "ods"])
    uploaded_file2 = st.file_uploader("Загрузите вторую таблицу (для сравнения)", type=["xlsx", "ods"])

    if uploaded_file1 and uploaded_file2:
        filetype1 = uploaded_file1.name.split(".")[-1].lower()
        filetype2 = uploaded_file2.name.split(".")[-1].lower()

        if (filetype1 == "ods" or filetype2 == "ods") and not odfpy_available:
            st.error("❌ Вы загрузили файл .ods, но библиотека 'odfpy' не установлена. Установите её: pip install odfpy")
            return

        try:
            df1 = pd.read_excel(uploaded_file1, engine="odf" if filetype1 == "ods" else None)
            df2 = pd.read_excel(uploaded_file2, engine="odf" if filetype2 == "ods" else None)
        except Exception as e:
            st.error(f"❌ Ошибка при чтении файлов: {e}")
            return

        st.success("✅ Таблицы загружены")
        st.write("Выберите столбец для поиска совпадений:")
        common_columns = list(set(df1.columns) & set(df2.columns))
        if not common_columns:
            st.error("Нет общих столбцов для сравнения.")
            return
        key_column = st.selectbox("🔑 Ключевой столбец:", common_columns)
        if st.button("Сравнить"):
            result_df = compare_tables(df1, df2, key_column)
            st.write("📋 Результаты сравнения:")
            st.dataframe(result_df)

            # Экспорт с цветами
            excel_data = generate_excel_with_colors(result_df)
            st.download_button(
                label="📥 Скачать результат в Excel (цветной)",
                data=excel_data,
                file_name="comparison_result_colored.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

main()